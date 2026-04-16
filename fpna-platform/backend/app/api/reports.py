"""
Reports API — real Excel export endpoints using openpyxl.
Endpoints: budget plan, variance analysis, baseline comparison, ad-hoc.
"""
import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from app.database import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/reports", tags=["Reports"])

# ── Styling helpers ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
TOTAL_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="1E293B", size=10)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
NUM_FMT = '#,##0.00'
PCT_FMT = '0.00"%"'


def _stream_workbook(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _apply_header_row(ws, row: int, cols: list):
    for i, (col_letter, title, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width


def _add_title_block(ws, title: str, subtitle: str, fiscal_year: int):
    ws.merge_cells("A1:P1")
    ws["A1"] = f"  {title} — FY {fiscal_year}"
    ws["A1"].font = Font(bold=True, size=14, color="1E293B")
    ws["A1"].fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:P2")
    ws["A2"] = f"  {subtitle}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=9, color="64748B", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18


# ── Budget Plan Export ─────────────────────────────────────────────────────

@router.get("/budget-plan/export")
def export_budget_plan(
    fiscal_year: int = Query(2026),
    db: Session = Depends(get_db)
):
    """Export full budget plan with monthly breakdown by department."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Plan"
    ws.freeze_panes = "D5"

    _add_title_block(ws, "Budget Plan", "Planned amounts by department, account, and month", fiscal_year)

    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = [
        ("A", "Department", 18), ("B", "Account Code", 14), ("C", "Account Name", 28),
        ("D", "Scenario", 12),
    ] + [(get_column_letter(5+i), m, 12) for i, m in enumerate(MONTHS)] + [
        (get_column_letter(17), "Annual Total", 16), (get_column_letter(18), "Status", 12)
    ]
    _apply_header_row(ws, 4, headers)

    try:
        q = """
            SELECT
                ISNULL(d.name_en, 'Unassigned') AS department,
                ISNULL(bpd.coa_code, '') AS account_code,
                ISNULL(bpd.coa_name, '') AS account_name,
                bp.status AS scenario,
                ISNULL(bpd.baseline_jan, 0), ISNULL(bpd.baseline_feb, 0), ISNULL(bpd.baseline_mar, 0),
                ISNULL(bpd.baseline_apr, 0), ISNULL(bpd.baseline_may, 0), ISNULL(bpd.baseline_jun, 0),
                ISNULL(bpd.baseline_jul, 0), ISNULL(bpd.baseline_aug, 0), ISNULL(bpd.baseline_sep, 0),
                ISNULL(bpd.baseline_oct, 0), ISNULL(bpd.baseline_nov, 0), ISNULL(bpd.baseline_dec, 0),
                ISNULL(bpd.baseline_total, 0) AS annual_total,
                bp.status
            FROM budget_plan_details bpd
            JOIN budget_plan_groups bpg ON bpd.group_id = bpg.id
            JOIN budget_plans bp ON bpg.plan_id = bp.id
            LEFT JOIN departments d ON bp.department_id = d.id
            WHERE bp.fiscal_year = :fy
            ORDER BY department, account_code
        """
        rows = db.execute(text(q), {"fy": fiscal_year}).fetchall()
    except Exception as e:
        rows = []
        logger.warning("Budget plan export query error: %s", e)

    row_num = 5
    dept_totals: dict = {}
    grand_annual = 0.0

    for i, row in enumerate(rows):
        dept = row[0]
        fill = ALT_FILL if i % 2 == 0 else None
        annual = float(row[16] or 0)
        grand_annual += annual
        dept_totals[dept] = dept_totals.get(dept, 0) + annual

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if col_idx >= 5 and col_idx <= 17:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    # Grand total row
    ws.cell(row=row_num, column=1, value="GRAND TOTAL").font = TOTAL_FONT
    ws.cell(row=row_num, column=1).fill = TOTAL_FILL
    # Sum columns 5-16 (monthly)
    for c in range(5, 18):
        cell = ws.cell(row=row_num, column=c)
        cell.value = f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{row_num-1})"
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.number_format = NUM_FMT
        cell.alignment = Alignment(horizontal="right")
    for c in range(1, 19):
        ws.cell(row=row_num, column=c).border = THIN_BORDER

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Department"
    ws2["B1"] = "Annual Total"
    ws2["A1"].font = HEADER_FONT
    ws2["B1"].font = HEADER_FONT
    ws2["A1"].fill = HEADER_FILL
    ws2["B1"].fill = HEADER_FILL
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 18
    for r, (dept, total) in enumerate(dept_totals.items(), 2):
        ws2.cell(row=r, column=1, value=dept)
        cell = ws2.cell(row=r, column=2, value=total)
        cell.number_format = NUM_FMT

    return _stream_workbook(wb, f"budget_plan_FY{fiscal_year}_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ── Variance Analysis Export ───────────────────────────────────────────────

@router.get("/variance/export")
def export_variance(
    fiscal_year: int = Query(2026),
    db: Session = Depends(get_db)
):
    """Export plan vs actual variance with % deviation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variance Analysis"
    ws.freeze_panes = "C5"

    _add_title_block(ws, "Variance Analysis", "Plan vs Actuals — deviation and % variance by account", fiscal_year)

    headers = [
        ("A", "Department", 18), ("B", "Account Code", 14), ("C", "Account", 28),
        ("D", "Plan (Annual)", 16), ("E", "Actual (YTD)", 16),
        ("F", "Variance $", 16), ("G", "Variance %", 12), ("H", "Status", 12),
    ]
    _apply_header_row(ws, 4, headers)

    try:
        q = """
            SELECT
                ISNULL(abf.department_name, 'Unassigned'),
                ISNULL(abf.coa_code, ''),
                ISNULL(abf.coa_name, ''),
                ISNULL(SUM(abf.baseline_amount), 0) AS plan_amt,
                ISNULL(SUM(abf.adjusted_amount), 0) AS actual_amt
            FROM approved_budget_fact abf
            WHERE abf.fiscal_year = :fy
            GROUP BY abf.department_name, abf.coa_code, abf.coa_name
            ORDER BY 1, 2
        """
        rows = db.execute(text(q), {"fy": fiscal_year}).fetchall()
    except Exception as e:
        rows = []
        logger.warning("Variance export query error: %s", e)

    row_num = 5
    for i, row in enumerate(rows):
        dept, code, name = row[0], row[1], row[2]
        plan = float(row[3] or 0)
        actual = float(row[4] or 0)
        variance = actual - plan
        variance_pct = (variance / plan * 100) if plan else 0
        status = "Over" if variance > plan * 0.1 else ("Under" if variance < -plan * 0.1 else "On Track")

        fill = ALT_FILL if i % 2 == 0 else None
        data = [dept, code, name, plan, actual, variance, variance_pct / 100, status]
        fmts = [None, None, None, NUM_FMT, NUM_FMT, NUM_FMT, '0.00%', None]

        for col_idx, (val, fmt) in enumerate(zip(data, fmts), 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
            # Color variance column
            if col_idx == 6:
                cell.font = Font(size=9, color="059669" if variance >= 0 else "DC2626")
        row_num += 1

    # Total row
    for c, label in [(1, "TOTAL"), (4, None), (5, None), (6, None)]:
        cell = ws.cell(row=row_num, column=c)
        if label:
            cell.value = label
        else:
            cell.value = f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{row_num-1})"
            cell.number_format = NUM_FMT
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    return _stream_workbook(wb, f"variance_FY{fiscal_year}_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ── Baseline Comparison Export ─────────────────────────────────────────────

@router.get("/baseline/export")
def export_baseline_comparison(
    fiscal_year: int = Query(2026),
    db: Session = Depends(get_db)
):
    """Export baseline vs planned comparison by account."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Baseline vs Plan"

    _add_title_block(ws, "Baseline Comparison", "AI-generated baseline vs approved plan", fiscal_year)

    headers = [
        ("A", "Account Code", 14), ("B", "Account Name", 30),
        ("C", "Baseline", 16), ("D", "Plan", 16),
        ("E", "Delta $", 16), ("F", "Delta %", 12),
    ]
    _apply_header_row(ws, 4, headers)

    try:
        q = """
            SELECT
                ISNULL(abf.coa_code, ''),
                ISNULL(abf.coa_name, abf.coa_code),
                ISNULL(SUM(abf.baseline_amount), 0) AS baseline,
                ISNULL(SUM(abf.adjusted_amount), 0) AS plan_total
            FROM approved_budget_fact abf
            WHERE abf.fiscal_year = :fy
            GROUP BY abf.coa_code, abf.coa_name
            ORDER BY abf.coa_code
        """
        rows = db.execute(text(q), {"fy": fiscal_year}).fetchall()
    except Exception as e:
        rows = []
        logger.warning("Baseline export query error: %s", e)

    row_num = 5
    for i, row in enumerate(rows):
        baseline = float(row[2] or 0)
        plan = float(row[3] or 0)
        delta = plan - baseline
        delta_pct = (delta / baseline * 100) if baseline else 0
        fill = ALT_FILL if i % 2 == 0 else None

        for col_idx, (val, fmt) in enumerate(zip(
            [row[0], row[1], baseline, plan, delta, delta_pct / 100],
            [None, None, NUM_FMT, NUM_FMT, NUM_FMT, '0.00%']
        ), 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    return _stream_workbook(wb, f"baseline_comparison_FY{fiscal_year}_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ── Ad-hoc Export ──────────────────────────────────────────────────────────

from pydantic import BaseModel

class AdhocExportRequest(BaseModel):
    fiscal_year: int = 2026
    dataset: str = "Planned Budgets"   # Planned Budgets | Actuals | Baselines | Variance
    group_by: str = "Department"       # Department | Account | Business Unit | Month
    period: str = "Full Year"          # Full Year | Q1..Q4 | YTD


@router.post("/adhoc/export")
def export_adhoc(request: AdhocExportRequest, db: Session = Depends(get_db)):
    """Ad-hoc query export — dataset + group by + period filter."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ad-hoc Export"

    _add_title_block(
        ws,
        f"Ad-hoc: {request.dataset}",
        f"Grouped by {request.group_by} | Period: {request.period}",
        request.fiscal_year
    )

    # Determine month range from period
    month_ranges = {
        "Full Year": (1, 12), "Q1": (1, 3), "Q2": (4, 6),
        "Q3": (7, 9), "Q4": (10, 12), "YTD": (1, datetime.now().month)
    }
    m_start, m_end = month_ranges.get(request.period, (1, 12))

    headers = [("A", request.group_by, 24), ("B", "Total Amount", 18), ("C", "% of Total", 12)]
    _apply_header_row(ws, 4, headers)

    MONTH_COL_NAMES = [
        "baseline_jan", "baseline_feb", "baseline_mar", "baseline_apr",
        "baseline_may", "baseline_jun", "baseline_jul", "baseline_aug",
        "baseline_sep", "baseline_oct", "baseline_nov", "baseline_dec"
    ]
    try:
        if request.dataset == "Planned Budgets":
            month_cols = " + ".join([f"ISNULL(bpd.{MONTH_COL_NAMES[m-1]}, 0)" for m in range(m_start, m_end + 1)])
            if request.group_by == "Department":
                q = f"""
                    SELECT ISNULL(d.name_en, 'Unassigned'), SUM({month_cols})
                    FROM budget_plan_details bpd
                    JOIN budget_plan_groups bpg ON bpd.group_id = bpg.id
                    JOIN budget_plans bp ON bpg.plan_id = bp.id
                    LEFT JOIN departments d ON bp.department_id = d.id
                    WHERE bp.fiscal_year = :fy GROUP BY d.name_en ORDER BY 2 DESC
                """
            elif request.group_by == "Account":
                q = f"""
                    SELECT ISNULL(bpg.coa_code, 'N/A'), SUM({month_cols})
                    FROM budget_plan_details bpd
                    JOIN budget_plan_groups bpg ON bpd.group_id = bpg.id
                    JOIN budget_plans bp ON bpg.plan_id = bp.id
                    WHERE bp.fiscal_year = :fy GROUP BY bpg.coa_code ORDER BY 2 DESC
                """
            else:
                q = f"""
                    SELECT ISNULL(bp.scenario, 'BASE'), SUM({month_cols})
                    FROM budget_plan_details bpd
                    JOIN budget_plan_groups bpg ON bpd.group_id = bpg.id
                    JOIN budget_plans bp ON bpg.plan_id = bp.id
                    WHERE bp.fiscal_year = :fy GROUP BY bp.scenario ORDER BY 2 DESC
                """
            rows = db.execute(text(q), {"fy": request.fiscal_year}).fetchall()
        elif request.dataset == "Baselines":
            q = """
                SELECT coa_code, SUM(baseline_amount) as baseline_total
                FROM approved_budget_fact
                WHERE fiscal_year = :fy
                GROUP BY coa_code
                ORDER BY 2 DESC
            """
            rows = db.execute(text(q), {"fy": request.fiscal_year}).fetchall()
        else:
            rows = []
    except Exception as e:
        rows = []
        logger.warning("Adhoc export error: %s", e)

    total = sum(float(r[1] or 0) for r in rows)
    row_num = 5
    for i, row in enumerate(rows):
        amount = float(row[1] or 0)
        pct = (amount / total) if total else 0
        fill = ALT_FILL if i % 2 == 0 else None
        for col_idx, (val, fmt) in enumerate(zip([row[0], amount, pct], [None, NUM_FMT, '0.00%']), 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    return _stream_workbook(
        wb,
        f"adhoc_{request.dataset.replace(' ', '_')}_{request.fiscal_year}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
