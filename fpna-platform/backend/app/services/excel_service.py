# backend/app/services/excel_service.py
"""
Excel processing service for FPNA data uploads

Supports two upload types:
1. Balance Snapshots - Monthly balance data for baseline calculation
2. Budget Planned - Ready budget plans for approval workflow
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict, List, Any, Optional
from datetime import datetime, date
from decimal import Decimal
import os
import uuid


class ExcelUploadType:
    """Upload type constants"""
    BALANCE_SNAPSHOT = "balance_snapshot"
    BUDGET_PLANNED = "budget_planned"


class ExcelProcessor:
    """Process Excel files for FPNA data uploads"""
    
    # Column mappings for balance snapshot uploads
    BALANCE_SNAPSHOT_COLUMNS = {
        'account_code': ['Account Code', 'account_code', 'AccountCode', 'ACCOUNT_CODE', 'Счет', 'Код счета'],
        'snapshot_date': ['Snapshot Date', 'snapshot_date', 'Date', 'date', 'Дата', 'Period'],
        'currency': ['Currency', 'currency', 'Валюта', 'CCY'],
        'balance': ['Balance', 'balance', 'Amount', 'Остаток', 'Сумма'],
        'balance_uzs': ['Balance UZS', 'balance_uzs', 'Amount UZS', 'Остаток UZS'],
        'fx_rate': ['FX Rate', 'fx_rate', 'Exchange Rate', 'Курс'],
        'branch_code': ['Branch', 'branch_code', 'Branch Code', 'Филиал'],
    }
    
    # Column mappings for budget planned uploads
    BUDGET_PLANNED_COLUMNS = {
        'account_code': ['Account Code', 'account_code', 'AccountCode', 'ACCOUNT_CODE', 'Счет'],
        'department': ['Department', 'department', 'Dept', 'Отдел'],
        'branch': ['Branch', 'branch', 'Branch Code', 'Филиал'],
        'currency': ['Currency', 'currency', 'Валюта', 'CCY'],
        'jan': ['Jan', 'January', 'jan', '01', 'Январь', 'M01'],
        'feb': ['Feb', 'February', 'feb', '02', 'Февраль', 'M02'],
        'mar': ['Mar', 'March', 'mar', '03', 'Март', 'M03'],
        'apr': ['Apr', 'April', 'apr', '04', 'Апрель', 'M04'],
        'may': ['May', 'may', '05', 'Май', 'M05'],
        'jun': ['Jun', 'June', 'jun', '06', 'Июнь', 'M06'],
        'jul': ['Jul', 'July', 'jul', '07', 'Июль', 'M07'],
        'aug': ['Aug', 'August', 'aug', '08', 'Август', 'M08'],
        'sep': ['Sep', 'September', 'sep', '09', 'Сентябрь', 'M09'],
        'oct': ['Oct', 'October', 'oct', '10', 'Октябрь', 'M10'],
        'nov': ['Nov', 'November', 'nov', '11', 'Ноябрь', 'M11'],
        'dec': ['Dec', 'December', 'dec', '12', 'Декабрь', 'M12'],
        'scenario': ['Scenario', 'scenario', 'Сценарий'],
        'notes': ['Notes', 'notes', 'Comments', 'Примечания'],
    }

    @staticmethod
    def _find_column(df_columns: List[str], possible_names: List[str]) -> Optional[str]:
        """Find matching column name from possible names"""
        df_cols_lower = {c.lower().strip(): c for c in df_columns}
        for name in possible_names:
            if name.lower().strip() in df_cols_lower:
                return df_cols_lower[name.lower().strip()]
        return None

    @staticmethod
    def _parse_date(value) -> Optional[date]:
        """Parse date from various formats"""
        if pd.isna(value):
            return None
        if isinstance(value, (datetime, date)):
            return value if isinstance(value, date) else value.date()
        try:
            return pd.to_datetime(value).date()
        except:
            return None

    @staticmethod
    def _parse_decimal(value, default: float = 0) -> Decimal:
        """Parse decimal value safely"""
        if pd.isna(value):
            return Decimal(str(default))
        try:
            return Decimal(str(float(value)))
        except:
            return Decimal(str(default))

    @classmethod
    def parse_balance_snapshot_excel(cls, file_path: str, fiscal_year: Optional[int] = None) -> Dict[str, Any]:
        """
        Parse Excel file for balance snapshot data
        
        Expected format - single sheet with columns:
        - Account Code (required)
        - Snapshot Date (required) - date of the balance
        - Currency (default: UZS)
        - Balance (required) - balance in original currency
        - Balance UZS (optional) - balance in UZS
        - FX Rate (optional, default: 1.0)
        - Branch (optional, default: ALL)
        
        Returns dict with:
        - records: list of balance snapshot records
        - summary: import statistics
        - import_batch_id: unique batch identifier
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Find required columns
            account_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['account_code'])
            date_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['snapshot_date'])
            balance_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['balance'])
            
            if not account_col:
                raise ValueError("Missing required column: Account Code")
            if not date_col:
                raise ValueError("Missing required column: Snapshot Date")
            if not balance_col:
                raise ValueError("Missing required column: Balance")
            
            # Find optional columns
            currency_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['currency'])
            balance_uzs_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['balance_uzs'])
            fx_rate_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['fx_rate'])
            branch_col = cls._find_column(df.columns, cls.BALANCE_SNAPSHOT_COLUMNS['branch_code'])
            
            import_batch_id = f"EXCEL-SNAP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"
            records = []
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    account_code = str(row[account_col]).strip()
                    if not account_code or account_code == 'nan':
                        continue
                    
                    snapshot_date = cls._parse_date(row[date_col])
                    if not snapshot_date:
                        errors.append(f"Row {idx + 2}: Invalid date")
                        continue
                    
                    balance = cls._parse_decimal(row[balance_col])
                    currency = str(row[currency_col]).strip().upper() if currency_col and pd.notna(row[currency_col]) else 'UZS'
                    
                    # Calculate balance_uzs if not provided
                    if balance_uzs_col and pd.notna(row[balance_uzs_col]):
                        balance_uzs = cls._parse_decimal(row[balance_uzs_col])
                    else:
                        balance_uzs = balance if currency == 'UZS' else Decimal('0')
                    
                    fx_rate = cls._parse_decimal(row[fx_rate_col], 1.0) if fx_rate_col else Decimal('1.0')
                    branch_code = str(row[branch_col]).strip() if branch_col and pd.notna(row[branch_col]) else 'ALL'
                    
                    record = {
                        'account_code': account_code,
                        'snapshot_date': snapshot_date,
                        'currency': currency,
                        'balance': float(balance),
                        'balance_uzs': float(balance_uzs),
                        'fx_rate': float(fx_rate),
                        'data_source': 'EXCEL_UPLOAD',
                        'import_batch_id': import_batch_id,
                        'is_validated': False,
                    }
                    records.append(record)
                    
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")
            
            if not records:
                raise ValueError("No valid records found in Excel file")
            
            # Extract date range
            dates = [r['snapshot_date'] for r in records]
            min_date = min(dates)
            max_date = max(dates)
            
            return {
                'upload_type': ExcelUploadType.BALANCE_SNAPSHOT,
                'records': records,
                'import_batch_id': import_batch_id,
                'summary': {
                    'total_rows': len(df),
                    'valid_records': len(records),
                    'error_count': len(errors),
                    'errors': errors[:20],  # First 20 errors
                    'date_range': {
                        'start': min_date.isoformat(),
                        'end': max_date.isoformat(),
                    },
                    'unique_accounts': len(set(r['account_code'] for r in records)),
                    'currencies': list(set(r['currency'] for r in records)),
                }
            }
            
        except Exception as e:
            raise ValueError(f"Error parsing balance snapshot Excel: {str(e)}")

    @classmethod
    def parse_budget_planned_excel(cls, file_path: str, fiscal_year: int) -> Dict[str, Any]:
        """
        Parse Excel file for budget planned data
        
        Expected format - single sheet with columns:
        - Account Code (required)
        - Department (optional)
        - Branch (optional)
        - Currency (default: UZS)
        - Jan, Feb, Mar, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec (monthly amounts)
        - Scenario (optional, default: BASE)
        - Notes (optional)
        
        Returns dict with:
        - records: list of budget planned records
        - summary: import statistics
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Find required columns
            account_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['account_code'])
            if not account_col:
                raise ValueError("Missing required column: Account Code")
            
            # Find optional columns
            dept_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['department'])
            branch_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['branch'])
            currency_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['currency'])
            scenario_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['scenario'])
            notes_col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS['notes'])
            
            # Find monthly columns
            month_cols = {}
            for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']:
                col = cls._find_column(df.columns, cls.BUDGET_PLANNED_COLUMNS[month])
                month_cols[month] = col
            
            # Check if at least one month column exists
            if not any(month_cols.values()):
                raise ValueError("No monthly amount columns found (Jan-Dec)")
            
            records = []
            errors = []
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            
            for idx, row in df.iterrows():
                try:
                    account_code = str(row[account_col]).strip()
                    if not account_code or account_code == 'nan':
                        continue
                    
                    # Generate unique budget code
                    budget_code = f"BPL-{fiscal_year}-{account_code}-{timestamp}-{idx}"
                    
                    department = str(row[dept_col]).strip() if dept_col and pd.notna(row[dept_col]) else None
                    branch = str(row[branch_col]).strip() if branch_col and pd.notna(row[branch_col]) else None
                    currency = str(row[currency_col]).strip().upper() if currency_col and pd.notna(row[currency_col]) else 'UZS'
                    scenario = str(row[scenario_col]).strip().upper() if scenario_col and pd.notna(row[scenario_col]) else 'BASE'
                    notes = str(row[notes_col]).strip() if notes_col and pd.notna(row[notes_col]) else None
                    
                    # Parse monthly values
                    monthly = {}
                    annual_total = Decimal('0')
                    for month, col in month_cols.items():
                        if col:
                            val = cls._parse_decimal(row[col])
                            monthly[month] = float(val)
                            annual_total += val
                        else:
                            monthly[month] = 0.0
                    
                    record = {
                        'budget_code': budget_code,
                        'fiscal_year': fiscal_year,
                        'account_code': account_code,
                        'department': department,
                        'branch': branch,
                        'currency': currency,
                        'scenario': scenario,
                        'notes': notes,
                        'annual_total': float(annual_total),
                        'annual_total_uzs': float(annual_total) if currency == 'UZS' else 0.0,
                        'status': 'DRAFT',
                        **monthly,
                    }
                    records.append(record)
                    
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")
            
            if not records:
                raise ValueError("No valid records found in Excel file")
            
            return {
                'upload_type': ExcelUploadType.BUDGET_PLANNED,
                'fiscal_year': fiscal_year,
                'records': records,
                'summary': {
                    'total_rows': len(df),
                    'valid_records': len(records),
                    'error_count': len(errors),
                    'errors': errors[:20],
                    'unique_accounts': len(set(r['account_code'] for r in records)),
                    'departments': list(set(r['department'] for r in records if r['department'])),
                    'branches': list(set(r['branch'] for r in records if r['branch'])),
                    'total_annual_amount': sum(r['annual_total'] for r in records),
                    'scenarios': list(set(r['scenario'] for r in records)),
                }
            }
            
        except Exception as e:
            raise ValueError(f"Error parsing budget planned Excel: {str(e)}")

    @staticmethod
    def create_balance_snapshot_template(file_path: str) -> str:
        """Create Excel template for balance snapshot upload"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Snapshots"
        
        # Headers
        headers = ['Account Code', 'Snapshot Date', 'Currency', 'Balance', 'Balance UZS', 'FX Rate', 'Branch']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['10101', '2025-01-31', 'UZS', 1500000000, 1500000000, 1.0, 'HQ'],
            ['10101', '2025-02-28', 'UZS', 1650000000, 1650000000, 1.0, 'HQ'],
            ['10201', '2025-01-31', 'USD', 125000, 1562500000, 12500, 'HQ'],
            ['10201', '2025-02-28', 'USD', 130000, 1625000000, 12500, 'HQ'],
            ['40101', '2025-01-31', 'UZS', 850000000, 850000000, 1.0, 'BRANCH01'],
            ['40101', '2025-02-28', 'UZS', 920000000, 920000000, 1.0, 'BRANCH01'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Add instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            "Balance Snapshot Upload Instructions",
            "",
            "Required Columns:",
            "- Account Code: GL account code (e.g., 10101)",
            "- Snapshot Date: End of month date (e.g., 2025-01-31)",
            "- Balance: Balance amount in original currency",
            "",
            "Optional Columns:",
            "- Currency: Currency code (default: UZS)",
            "- Balance UZS: Balance in UZS (auto-calculated if not provided)",
            "- FX Rate: Exchange rate to UZS (default: 1.0)",
            "- Branch: Branch code (default: ALL)",
            "",
            "Notes:",
            "- Each row represents one account's balance at a specific date",
            "- Upload multiple months at once for baseline calculation",
            "- Recommended: 12-36 months of historical data",
        ]
        for row_idx, text in enumerate(instructions, 1):
            ws_inst.cell(row=row_idx, column=1, value=text)
        ws_inst.column_dimensions['A'].width = 60
        
        wb.save(file_path)
        return file_path

    @staticmethod
    def create_budget_planned_template(file_path: str, fiscal_year: int = None) -> str:
        """Create Excel template for budget planned upload"""
        if fiscal_year is None:
            fiscal_year = datetime.now().year + 1
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget Plan"
        
        # Headers
        headers = ['Account Code', 'Department', 'Branch', 'Currency', 
                   'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
                   'Scenario', 'Notes']
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Sample data
        sample_data = [
            ['40101', 'Sales', 'HQ', 'UZS', 100000000, 105000000, 110000000, 115000000, 120000000, 125000000,
             130000000, 125000000, 120000000, 115000000, 110000000, 105000000, 'BASE', 'Revenue target'],
            ['50101', 'Operations', 'HQ', 'UZS', 40000000, 42000000, 44000000, 46000000, 48000000, 50000000,
             52000000, 50000000, 48000000, 46000000, 44000000, 42000000, 'BASE', 'COGS budget'],
            ['60101', 'HR', 'HQ', 'UZS', 25000000, 25000000, 25000000, 25000000, 25000000, 25000000,
             25000000, 25000000, 25000000, 25000000, 25000000, 25000000, 'BASE', 'Salaries'],
            ['60201', 'Admin', 'HQ', 'UZS', 8000000, 8000000, 8000000, 8000000, 8000000, 8000000,
             8000000, 8000000, 8000000, 8000000, 8000000, 8000000, 'BASE', 'Rent'],
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        for col in 'EFGHIJKLMNOP':
            ws.column_dimensions[col].width = 14
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 20
        
        # Add instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            f"Budget Plan Upload Instructions - Fiscal Year {fiscal_year}",
            "",
            "Required Columns:",
            "- Account Code: GL account code (e.g., 40101)",
            "",
            "Optional Columns:",
            "- Department: Department name",
            "- Branch: Branch code",
            "- Currency: Currency code (default: UZS)",
            "- Jan-Dec: Monthly budget amounts",
            "- Scenario: BASE, OPTIMISTIC, or PESSIMISTIC (default: BASE)",
            "- Notes: Additional comments",
            "",
            "Workflow:",
            "1. Upload creates budget records in DRAFT status",
            "2. Review and modify as needed",
            "3. Apply driver adjustments if required",
            "4. Submit for approval workflow",
            "5. After approval, export to DWH",
            "",
            "Notes:",
            "- Each row represents one account's annual budget",
            "- Monthly columns can be left empty (defaults to 0)",
            "- Multiple scenarios can be uploaded for same account",
        ]
        for row_idx, text in enumerate(instructions, 1):
            ws_inst.cell(row=row_idx, column=1, value=text)
        ws_inst.column_dimensions['A'].width = 60
        
        wb.save(file_path)
        return file_path

    @staticmethod
    def parse_budget_excel(file_path: str) -> Dict[str, Any]:
        """
        Legacy method - Parse Excel file for old budget format
        Kept for backward compatibility
        """
        try:
            xl_file = pd.ExcelFile(file_path)
            
            if 'Header' not in xl_file.sheet_names or 'LineItems' not in xl_file.sheet_names:
                raise ValueError("Excel must contain 'Header' and 'LineItems' sheets")
            
            header_df = pd.read_excel(file_path, sheet_name='Header')
            header_data = {}
            for _, row in header_df.iterrows():
                if 'Field' in row and 'Value' in row:
                    header_data[str(row['Field']).strip()] = row['Value']
            
            items_df = pd.read_excel(file_path, sheet_name='LineItems')
            items_df.columns = items_df.columns.str.strip()
            
            line_items = []
            for _, row in items_df.iterrows():
                item = {
                    'account_code': str(row.get('Account Code', '')).strip(),
                    'account_name': str(row.get('Account Name', '')).strip(),
                    'category': str(row.get('Category', '')).strip() if pd.notna(row.get('Category')) else None,
                    'month': int(row.get('Month')) if pd.notna(row.get('Month')) else None,
                    'quarter': int(row.get('Quarter')) if pd.notna(row.get('Quarter')) else None,
                    'year': int(row.get('Year')) if pd.notna(row.get('Year')) else None,
                    'amount': float(row.get('Amount', 0)),
                    'quantity': float(row.get('Quantity')) if pd.notna(row.get('Quantity')) else None,
                    'unit_price': float(row.get('Unit Price')) if pd.notna(row.get('Unit Price')) else None,
                    'notes': str(row.get('Notes', '')).strip() if pd.notna(row.get('Notes')) else None,
                }
                
                if not item['account_code'] or not item['account_name']:
                    continue
                
                line_items.append(item)
            
            if not line_items:
                raise ValueError("No valid line items found in Excel file")
            
            total_amount = sum(item['amount'] for item in line_items)
            
            return {
                'header': {
                    'fiscal_year': int(header_data.get('Fiscal Year', datetime.now().year)),
                    'department': str(header_data.get('Department', '')),
                    'branch': str(header_data.get('Branch', '')),
                    'description': str(header_data.get('Description', '')),
                    'currency': str(header_data.get('Currency', 'USD')),
                },
                'line_items': line_items,
                'total_amount': total_amount,
                'summary': {
                    'total_items': len(line_items),
                    'total_amount': total_amount,
                    'categories': list(set(item['category'] for item in line_items if item['category']))
                }
            }
            
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    @staticmethod
    def create_template(file_path: str):
        """Create Excel template for budget upload"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create Header sheet
        ws_header = wb.active
        ws_header.title = "Header"
        ws_header['A1'] = "Field"
        ws_header['B1'] = "Value"
        ws_header['A2'] = "Fiscal Year"
        ws_header['B2'] = 2025
        ws_header['A3'] = "Department"
        ws_header['B3'] = "Finance"
        ws_header['A4'] = "Branch"
        ws_header['B4'] = "Head Office"
        ws_header['A5'] = "Description"
        ws_header['B5'] = "Annual Budget 2025"
        ws_header['A6'] = "Currency"
        ws_header['B6'] = "USD"
        
        # Style header sheet
        for cell in ws_header[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Create LineItems sheet
        ws_items = wb.create_sheet("LineItems")
        headers = [
            "Account Code", "Account Name", "Category", "Month", 
            "Quarter", "Year", "Amount", "Quantity", "Unit Price", "Notes"
        ]
        ws_items.append(headers)
        
        # Add sample data
        sample_data = [
            ["4000", "Revenue - Sales", "Revenue", 1, 1, 2025, 100000, 500, 200, "Q1 Sales Target"],
            ["4010", "Revenue - Services", "Revenue", 1, 1, 2025, 50000, None, None, "Consulting"],
            ["5000", "Cost of Goods Sold", "COGS", 1, 1, 2025, 40000, 500, 80, "Product Costs"],
            ["6000", "Salaries", "Operating Expenses", 1, 1, 2025, 60000, None, None, "Staff Salaries"],
            ["6100", "Rent", "Operating Expenses", 1, 1, 2025, 15000, None, None, "Office Rent"],
        ]
        
        for row in sample_data:
            ws_items.append(row)
        
        # Style LineItems sheet
        for cell in ws_items[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Adjust column widths
        for ws in [ws_header, ws_items]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(file_path)
        
        return file_path